import datetime as dt
import os
import sys
from glob import glob

import numpy as np
import openpyxl
import pandas as pd
import warnings
import xlrd

# TDT_SAMPLE_DELAY = 21  # TDT introduces a delay between stim and recording. This is the sample size of that delay
# PULSE_WIDTH_DELAY = .20e-3


def electrode_distances(log_path):
    wb = openpyxl.load_workbook(log_path)
    ws = wb.active
    num_header_cells = 1
    num_recordings = 0
    cols = tuple(ws.columns)
    target_col = "E"
    target_col_num = ord(target_col.lower()) - 96
    max_length = len(cols[target_col_num - 1])
    distances = []

    for row in range(1, max_length + 1):
        if ws.cell(row=row, column=target_col_num).value is None:
            num_recordings = row - 2
            break
    if ws.cell(row=max_length, column=target_col_num).value is not None:
        num_recordings = max_length - 1

    for i in range(num_recordings):
        distances.append(ws.cell(i + num_header_cells + 1, target_col_num).value)

    return distances


def check_make_dir(input_dir):
    if not os.path.isdir(input_dir):
        os.makedirs(input_dir)


def get_exp(input_dir):
    pass


def create_experimental_log(file_path):
    # First, we'll get all tdt tanks and store them in tdt_tanks
    tdt_tanks = []

    directories = [x[0] for x in os.walk(file_path)]

    # TDT tanks are a directory with a number of files. In particular any stream data is contained in a *.tev file.
    # Let's use that as in indicator of a complete tdt tank.
    tev_search = "/*.tev"
    for d in directories:
        tev_files = glob(d + tev_search)
        # Append
        if len(tev_files) > 0:
            tdt_tanks.append(d)

    wb = openpyxl.Workbook()
    ws = wb.active

    # We'll create some headers in the log
    headers = ["Location", "Stimulation Contact", "Stimulation Description", "Experimental Condition",
               "Distances: Recording to Stimulating Electrode (cm)"]
    for idx, val in enumerate(headers):
        _cell = ws.cell(row=1, column=idx + 1, value=val)
        _cell.font = openpyxl.styles.Font(bold=True)

    for idx, val in enumerate(tdt_tanks):
        ws.cell(row=idx + 2, column=1, value=val)

    # This next bit of code "autosizes" the cell widths
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = value

    experiment_log_path = file_path + "/Experimental Log.xlsx"
    wb.save(filename=experiment_log_path)
    return experiment_log_path


def matlab2datetime(matlab_datenum):
    day = dt.datetime.fromordinal(int(matlab_datenum))
    dayfrac = dt.timedelta(days=matlab_datenum % 1) - dt.timedelta(days=366)
    return day + dayfrac
